import openpyxl
import os

def append_row_to_excel(filename, row_data):
    """
    Додає рядок до першого аркуша Excel-файлу (.xlsx).
    row_data: список значень для рядка.
    """
    if not os.path.exists(filename):
        return "Файл Excel не знайдено!"
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws.append(row_data)
        wb.save(filename)
        return "Рядок додано до Excel-документу."
    except Exception as e:
        return f"Не вдалося змінити Excel-документ: {e}"

def read_excel_column(filename, col_idx=1):
    """
    Зчитує всі значення з вказаної колонки (індексація з 1).
    """
    if not os.path.exists(filename):
        return []
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        return [cell.value for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True)][0]
    except Exception as e:
        return []